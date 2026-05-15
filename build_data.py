"""Preprocess all CSVs + Excel into a single data.js for the static site."""
import json
import csv
from openpyxl import load_workbook
from pathlib import Path

UPLOADS = Path("/mnt/user-data/uploads")
OUT = Path("/home/claude/site/js/data.js")


def read_csv_names(path, has_header=True, name_col=0):
    """Read a list of names from a CSV file."""
    names = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        if has_header:
            next(reader, None)
        for row in reader:
            if not row:
                continue
            n = (row[name_col] or "").strip()
            # remove trailing commas and whitespace
            n = n.rstrip(",").strip()
            if n:
                names.append(n)
    return names


def read_simple_list(path):
    """Read one-name-per-line file (no header, no comma)."""
    names = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            n = line.strip().rstrip(",").strip()
            if n:
                names.append(n)
    return names


def read_wids(path):
    """Read WiDS Excel into list of {name, project} dicts."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # header row: Name, Roll Number, Project Name
    entries = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        name = (str(r[0]) or "").strip()
        project = (str(r[2]) if r[2] else "").strip()
        if name and project:
            entries.append({"name": name, "project": project})
    return entries


def read_mentor(path):
    """Read mentor CSV: Name, Project Name."""
    entries = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = (row.get("Name") or "").strip()
            project = (row.get("Project Name") or "").strip()
            if name and project:
                entries.append({"name": name, "project": project})
    return entries


# Load all
cv_basic = read_csv_names(UPLOADS / "Deep_Learning_Bootcamp-CV-List_-_Basic_Track.csv")
cv_advanced = read_csv_names(UPLOADS / "Deep_Learning_Bootcamp-CV-List_-_Advance_Track.csv")
nlp_basic = read_simple_list(UPLOADS / "NLP_-_NLP_Basic.csv")
nlp_advanced = read_simple_list(UPLOADS / "NLP_-_NLP_Advanced.csv")
wids = read_wids(UPLOADS / "WiDS_Qualified_2025.xlsx")
mentors = read_mentor(UPLOADS / "WiDS_5_0_Mentor_Form___Responses__-_Sheet1.csv")

# Deduplicate while preserving first occurrence (uses normalized name as key)
def dedupe(names):
    seen = set()
    out = []
    for n in names:
        key = " ".join(n.lower().split())
        if key not in seen:
            seen.add(key)
            out.append(n)
    return out

cv_basic = dedupe(cv_basic)
cv_advanced = dedupe(cv_advanced)
nlp_basic = dedupe(nlp_basic)
nlp_advanced = dedupe(nlp_advanced)

# Dedupe WiDS: same person + same project = one entry
def dedupe_wids(entries):
    seen = set()
    out = []
    for e in entries:
        key = (" ".join(e["name"].lower().split()), " ".join(e["project"].lower().split()))
        if key not in seen:
            seen.add(key)
            out.append(e)
    return out

wids = dedupe_wids(wids)
mentors = dedupe_wids(mentors)

print(f"CV Basic: {len(cv_basic)}")
print(f"CV Advanced: {len(cv_advanced)}")
print(f"NLP Basic: {len(nlp_basic)}")
print(f"NLP Advanced: {len(nlp_advanced)}")
print(f"WiDS entries: {len(wids)}")
print(f"WiDS unique names: {len(set(e['name'].lower().strip() for e in wids))}")
print(f"Mentor entries: {len(mentors)}")

data = {
    "bootcamp": {
        "cv_basic": cv_basic,
        "cv_advanced": cv_advanced,
        "nlp_basic": nlp_basic,
        "nlp_advanced": nlp_advanced,
    },
    "wids": wids,
    "mentors": mentors,
}

OUT.parent.mkdir(parents=True, exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("// Auto-generated. Do not edit by hand.\n")
    f.write("window.CERT_DATA = ")
    json.dump(data, f, ensure_ascii=False, indent=2)
    f.write(";\n")

print(f"\nWrote {OUT}")
print(f"Size: {OUT.stat().st_size / 1024:.1f} KB")
